from openpyxl import Workbook, load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart ,LineChart, Reference,AreaChart,PieChart,RadarChart,BubbleChart

wb = Workbook()
ws = wb.active

pokemon_data = [
    {
        "Name": "妙蛙種子",
        "HP": 45,
        "Attack": 49,
        "Defense": 49,
        "SpecialAtk": 65,
        "SpecialDef": 65,
        "Speed": 45,
    },
    {
        "Name": "妙蛙草",
        "HP": 60,
        "Attack": 62,
        "Defense": 63,
        "SpecialAtk": 80,
        "SpecialDef": 80,
        "Speed": 60,
    },
    {
        "Name": "妙蛙花",
        "HP": 80,
        "Attack": 82,
        "Defense": 83,
        "SpecialAtk": 100,
        "SpecialDef": 100,
        "Speed": 80,
    },
    {
        "Name": "小火龍",
        "HP": 39,
        "Attack": 52,
        "Defense": 43,
        "SpecialAtk": 60,
        "SpecialDef": 50,
        "Speed": 65,
    },
    {
        "Name": "火恐龍",
        "HP": 58,
        "Attack": 64,
        "Defense": 58,
        "SpecialAtk": 80,
        "SpecialDef": 65,
        "Speed": 80,
    },
    {
        "Name": "噴火龍",
        "HP": 78,
        "Attack": 84,
        "Defense": 78,
        "SpecialAtk": 109,
        "SpecialDef": 85,
        "Speed": 100,
    },
    {
        "Name": "傑尼龜",
        "HP": 44,
        "Attack": 48,
        "Defense": 65,
        "SpecialAtk": 50,
        "SpecialDef": 64,
        "Speed": 43,
    },
    {
        "Name": "卡咪龜",
        "HP": 59,
        "Attack": 63,
        "Defense": 80,
        "SpecialAtk": 65,
        "SpecialDef": 80,
        "Speed": 58,
    },
    {
        "Name": "水箭龜",
        "HP": 79,
        "Attack": 83,
        "Defense": 100,
        "SpecialAtk": 85,
        "SpecialDef": 105,
        "Speed": 78,
    },
]
title = ["name", "HP", "Attack", "Defense", "SpecialAtk", "SpecialDef", "Speed"]

ws.append(title)
for pokemon in pokemon_data:
    ws.append(list(pokemon.values()))

for col in range(2,8):
    char = get_column_letter(col)
    ws[char + "11"] = f"=SUM({char}2:{char}10)"
    ws[char + "12"] = f"=AVERAGE({char}2:{char}10)"

wb.save("pokemon.xlsx")

data = Reference(ws,min_col=2,max_col=2,min_row=1,max_row=10)
label = Reference(ws,min_col=1,max_col=1,min_row=2,max_row=10)
chart = LineChart()
chart.style = 13
chart.title = "pokemon-HP"
chart.y_axis.title = "HP"
chart.x_axis.title = "Name"
chart.add_data(data, titles_from_data=True)
chart.set_categories(label)
ws.add_chart(chart,"I2")
wb.save("pokemon.xlsx")
# wb = Workbook()
# wb.create_sheet("name")
# ws = wb.active
# ws.title = "my sheet"
# ws["A1"].value = 123
# ws["A2"].value = 456
# ws["B1"].value = "abc"
# ws["B2"].value = "def"
# wb.save("excel.xlsx")



# wb = load_workbook("excel.xlsx")
# ws = wb.active
# for row in range(1, ws.max_row+1):
#     for col in range(1, ws.max_column+1):
#         char = get_column_letter(col)
#         print(ws[char + str(row)].value)




# ws = wb.create_chartsheet("name")
# ws = wb["name"]
# print(ws)
# print(wb.sheetnames)
# wb.save("excel.xlsx")

# wb = load_workbook('path')
# ws = wb.active

# print (ws)
# print(ws["b3"])
# print(ws["d2"].value)